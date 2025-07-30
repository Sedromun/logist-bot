import os
import datetime
import pandas as pd
from aiogram import Bot, Dispatcher, types
from aiogram.types import Message, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from aiogram.enums import ContentType
from aiogram import F
import asyncio
import openpyxl
import re

import logging
logging.basicConfig(level=logging.INFO)

API_TOKEN = os.getenv('API_TOKEN')

bot = Bot(token=API_TOKEN)
dp = Dispatcher()

# Простое хранилище для выбора пользователя (user_id: {selection, results})
user_selections = {}

def normalize_rc_name(name):
    name = str(name).replace('"', '').replace("'", '')
    name = name.lower()
    name = re.sub(r'^(ао\s+тандер\s+)', '', name)  # убираем префикс в начале
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'\s*,\s*', ',', name)
    name = re.sub(r'\s*\(\s*', '(', name)
    name = re.sub(r'\s*\)\s*', ')', name)
    name = name.strip()
    return name

def is_green(cell):
    fill = cell.fill
    if fill is not None and hasattr(fill, 'fgColor') and fill.fgColor is not None:
        fgColor = fill.fgColor
        color = None
        if hasattr(fgColor, 'rgb') and fgColor.rgb is not None:
            color = fgColor.rgb
        elif hasattr(fgColor, 'type') and fgColor.type == 'rgb' and hasattr(fgColor, 'value'):
            color = fgColor.value
        if color is not None and isinstance(color, str):
            return color.upper() in ['FF92D050', '92D050']
    return False

def has_green_in_last_4_days(ws, row_idx, date_cols, plan_date):
    # Берём 4 дня до планируемого: plan_date-1, plan_date-2, plan_date-3, plan_date-4
    days = [plan_date - datetime.timedelta(days=i) for i in range(1, 5)]
    header_row = 4  # pandas header=3, значит названия колонок в 4-й строке (индекс 4 для openpyxl)
    header_values = [cell.value for cell in ws[header_row]]
    col_indices = []
    for day in days:
        for col, date in date_cols:
            if date.date() == day:
                if col in header_values:
                    col_indices.append(header_values.index(col) + 1)
                break
    excel_row = row_idx + 5
    for col_idx in col_indices:
        cell = ws.cell(row=excel_row, column=col_idx)
        if is_green(cell):
            return True
    return False

def build_keyboard(rc_results, selection, rc_plan_dates):
    keyboard = []
    weekdays = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]
    for i, (rc, text) in enumerate(rc_results):
        mark = "✅" if selection[rc] else "❌"
        plan_date = rc_plan_dates[rc]
        plan_date_str = plan_date.strftime("%d.%m")
        weekday_str = weekdays[plan_date.weekday()]
        # Убираем 'АО "Тандер"' из начала названия РЦ
        rc_clean = rc.strip()
        if rc_clean.startswith('АО "Тандер"'):
            rc_clean = rc_clean.replace('АО "Тандер"', '', 1).strip()
        # Формируем текст кнопки без 'АО "Тандер"'
        if "оптимальное" in text:
            qty = text.split("оптимальное")[1].strip().split()[0]
            button_text = f"{rc_clean}: оптимальное {qty} {mark} на {plan_date_str} ({weekday_str})"
        elif "(сгорит" in text:
            parts = text.split(":", 1)[1].strip()
            qty = parts.split()[0]
            burn = parts.split("(сгорит")[1].split(")")[0].strip()
            button_text = f"{rc_clean}: {qty} (сгорит {burn}) {mark} на {plan_date_str} ({weekday_str})"
        else:
            button_text = f"{text} {mark} на {plan_date_str} ({weekday_str})"
        keyboard.append([InlineKeyboardButton(text=button_text, callback_data=f"toggle|{i}")])

    keyboard.append([InlineKeyboardButton(text=f"Сформировать запрос", callback_data="submit")])
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

@dp.message(F.content_type == ContentType.DOCUMENT)
async def handle_planning_file(message: Message):
    file_name = message.document.file_name
    if "запрос планирования" not in file_name.lower():
        await handle_registry_file(message)
        return  # Не наш файл — вызываем основной handler

    file = await bot.get_file(message.document.file_id)
    await bot.download(file, destination=file_name)
    df_plan = pd.read_excel(file_name)
    col_rc = "РЦ Возврата"
    col_date = "Плановая Дата"
    col_volume = "Объем к возврату (кратно 20 шт)"

    # Получаем исходные данные пользователя
    user_id = message.from_user.id
    state = user_selections.get(user_id)
    if not state:
        await message.answer("Нет исходного реестра для проверки лимитов. Сначала загрузите реестр!")
        os.remove(file_name)
        return
    date_cols = state["date_cols"]
    df = state["df"]
    # limits_dict и min/max
    limits_file = 'limits.csv'
    limits_df = pd.read_csv(limits_file)
    limits_dict = {normalize_rc_name(k): v for k, v in limits_df.set_index('РЦ').to_dict(orient='index').items()}

    results = []
    for idx, row in df_plan.iterrows():
        try:
            volume = int(row[col_volume])
        except Exception:
            continue
        rc = str(row[col_rc])
        plan_date = pd.to_datetime(str(row[col_date]), errors='coerce')
        # Отладочная информация
        print(f"Парсинг даты: '{row[col_date]}' -> {plan_date}")
        rc_norm = normalize_rc_name("АО \"Тандер\" " + rc)  # добавить префикс для поиска в реестре
        # Найти строку в исходном df
        df_idx = None
        for i, row_src in df.iterrows():
            rc_src = row_src[df.columns[3]]
            rc_src_norm = normalize_rc_name(rc_src)
            if rc_src_norm == rc_norm:
                df_idx = i
                break
        if df_idx is None or rc_norm not in limits_dict:
            continue
        min_val = limits_dict[rc_norm].get('min')
        max_val = limits_dict[rc_norm].get('max')
        # Определяем сгорает ли (ищем столбцы с датами: планируемый и следующий)
        plan_col = next((col for col, date in date_cols if date.date() == plan_date.date()), None)
        next_col = next((col for col, date in date_cols if date.date() == (plan_date + datetime.timedelta(days=1)).date()), None)
        if not plan_col or not next_col:
            continue
        try:
            val_plan = int(df.iloc[df_idx][plan_col])
            val_next = int(df.iloc[df_idx][next_col])
        except Exception:
            continue
        # Проверяем условия
        if val_next < val_plan:  # Сгорает
            if volume >= min_val:
                results.append(f"{rc} — {row[col_date]} — {volume}")
        else:  # Не сгорает
            if volume >= max_val:
                results.append(f"{rc} — {row[col_date]} — {volume}")

    if results:
        await message.answer("Нужно отгружать:\n" + "\n".join(results))
    else:
        await message.answer("Нет РЦ для отгрузки по этому файлу.")

    os.remove(file_name)

@dp.message(F.content_type == ContentType.DOCUMENT)
async def handle_registry_file(message: Message):
    # Сохраняем файл
    file = await bot.get_file(message.document.file_id)
    file_path = file.file_path
    file_name = message.document.file_name
    await bot.download(file, destination=file_name)

    # Читаем Excel
    df = pd.read_excel(file_name, sheet_name=0, header=3)

    # Читаем лимиты
    limits_file = 'limits.csv'
    if not os.path.exists(limits_file):
        await message.answer("Файл limits.csv не найден!")
        os.remove(file_name)
        return
    limits_df = pd.read_csv(limits_file)
    limits_dict = {normalize_rc_name(k): v for k, v in limits_df.set_index('РЦ').to_dict(orient='index').items()}

    # Находим столбцы с датами
    date_cols = []
    for col in df.columns:
        try:
            # Пробуем разные форматы дат
            date = pd.to_datetime(str(col), errors='coerce')
            if pd.notnull(date):
                date_cols.append((col, date))
        except Exception:
            continue

    # Определяем сегодняшнюю дату и нужные столбцы
    today = datetime.datetime.now().date()
    day_plus_3 = today + datetime.timedelta(days=3)
    day_plus_2 = today + datetime.timedelta(days=2)

    col3 = None
    col2 = None
    for col, date in date_cols:
        if date.date() == day_plus_3:
            col3 = col
        if date.date() == day_plus_2:
            col2 = col

    if not col3 or not col2:
        await message.answer("Не удалось найти нужные даты в файле.")
        os.remove(file_name)
        return
    else:
        await message.answer(f"формируем запрос на {day_plus_2} ({col2})")

    # Открываем Excel через openpyxl один раз
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Анализируем данные
    results = []
    rc_results = []  # Для хранения (rc, текст)
    for idx, row in df.iterrows():
        rc = row[df.columns[3]]  # Название РЦ из столбца D
        rc_norm = normalize_rc_name(rc)
        try:
            val3 = int(row[col3])
            val2 = int(row[col2])
        except Exception as e:
            continue
        min_val = None
        max_val = None
        if rc_norm in limits_dict:
            min_val = limits_dict[rc_norm].get('min')
            max_val = limits_dict[rc_norm].get('max')
        if min_val is None or max_val is None:
            continue  # пропускаем, если нет лимитов
        green_found = has_green_in_last_4_days(ws, idx, date_cols, day_plus_2)
        if ((val3 < val2 and val2 >= min_val) or val2 >= max_val) and not green_found:
            if val3 < val2 and val2 >= min_val:
                text = f"{rc}: {val2} (сгорит {val2-val3})"
                results.append(text)
                rc_results.append((rc, text))
            elif val2 >= max_val:
                text = f"{rc}: оптимальное {val2}"
                results.append(text)
                rc_results.append((rc, text))

    if results:
        user_id = message.from_user.id
        selection = {rc: True for rc, _ in rc_results}
        # Определяем даты планирования для каждого РЦ
        rc_plan_dates = {}
        today = datetime.datetime.now().date()
        weekday = today.weekday()
        if weekday == 4:  # Пятница
            day_sun = today + datetime.timedelta(days=2)
            day_mon = today + datetime.timedelta(days=3)
            day_tue = today + datetime.timedelta(days=4)
            # Находим названия столбцов для этих дат
            col_sun = col_mon = col_tue = None
            for col, date in date_cols:
                if date.date() == day_sun:
                    col_sun = col
                if date.date() == day_mon:
                    col_mon = col
                if date.date() == day_tue:
                    col_tue = col
            # Определяем дату планирования для каждого РЦ
            for rc, _ in rc_results:
                df_idx = None
                for idx, row in df.iterrows():
                    rc_df = row[df.columns[3]]
                    rc_df_clean = rc_df.strip()
                    if rc_df_clean.startswith('АО "Тандер"'):
                        rc_df_clean = rc_df_clean.replace('АО "Тандер"', '', 1).strip()
                    rc_clean = rc.strip()
                    if rc_clean.startswith('АО "Тандер"'):
                        rc_clean = rc_clean.replace('АО "Тандер"', '', 1).strip()
                    if rc_df_clean == rc_clean:
                        df_idx = idx
                        break
                plan_date = day_tue
                if df_idx is not None and col_sun and col_mon and col_tue:
                    row = df.iloc[df_idx]
                    try:
                        val_sun = int(row[col_sun])
                        val_mon = int(row[col_mon])
                        val_tue = int(row[col_tue])
                        if val_tue < val_mon:
                            plan_date = day_mon
                        if val_mon < val_sun:
                            plan_date = day_sun
                    except Exception:
                        pass
                rc_plan_dates[rc] = plan_date
        else:
            day_plus_2 = today + datetime.timedelta(days=2)
            for rc, _ in rc_results:
                rc_plan_dates[rc] = day_plus_2
        user_selections[user_id] = {"selection": selection, "rc_results": rc_results, "date_cols": date_cols, "df": df, "rc_plan_dates": rc_plan_dates}
        markup = build_keyboard(rc_results, selection, rc_plan_dates)
        await message.answer(
            "Выберите РЦ для запроса (нажмите, чтобы исключить):",
            reply_markup=markup
        )
    else:
        await message.answer("Нет подходящих РЦ по условиям.")

    os.remove(file_name)

@dp.callback_query(lambda c: c.data.startswith("toggle|"))
async def toggle_rc(query: CallbackQuery):
    user_id = query.from_user.id
    i = int(query.data.split("|", 1)[1])
    state = user_selections.get(user_id)
    if not state:
        await query.answer("Сессия устарела.")
        return
    rc_results = state["rc_results"]
    selection = state["selection"]
    rc_plan_dates = state["rc_plan_dates"]
    if i < 0 or i >= len(rc_results):
        await query.answer("Ошибка выбора.")
        return
    rc = rc_results[i][0]
    selection[rc] = not selection[rc]
    markup = build_keyboard(rc_results, selection, rc_plan_dates)
    await query.message.edit_reply_markup(reply_markup=markup)
    await query.answer()

import pandas as pd
from aiogram.types import FSInputFile

@dp.callback_query(lambda c: c.data == "submit")
async def submit_query(query: CallbackQuery):
    user_id = query.from_user.id
    state = user_selections.get(user_id)
    if not state:
        await query.answer("Сессия устарела.")
        return
    selection = state["selection"]
    rc_results = state["rc_results"]
    date_cols = state["date_cols"]
    df = state["df"]
    filtered = [text for rc, text in rc_results if selection[rc]]
    if not filtered:
        await query.message.answer("Вы не выбрали ни одного РЦ.")
        await query.answer()
        return

    # --- Формируем Excel ---
    columns = [
        "Плановая Дата",
        "ИНН КА",
        "НАИМЕНОВАНИЕ КА",
        "№ Договора",
        "РЦ Возврата",
        "КОЛ-ВО поддонов к возврату (Фактически отгруженных на РЦ) по данным КА",
        "КОЛ-ВО поддонов к возврату (Фактически отгруженных на РЦ) по данным АО \"Тандер\"",
        "Объем к возврату (кратно 20 шт)"
    ]
    today = datetime.datetime.now().date()
    weekday = today.weekday()  # 0=Пн, 4=Пт, 6=Вс
    rows = []
    if weekday == 4:  # Пятница
        day_sun = today + datetime.timedelta(days=2)
        day_mon = today + datetime.timedelta(days=3)
        day_tue = today + datetime.timedelta(days=4)
        # Находим названия столбцов для этих дат
        col_sun = col_mon = col_tue = None
        for col, date in date_cols:
            if date.date() == day_sun:
                col_sun = col
            if date.date() == day_mon:
                col_mon = col
            if date.date() == day_tue:
                col_tue = col
        for text in filtered:
            rc, rest = text.split(":", 1)
            rc_clean = rc.strip()
            if rc_clean.startswith('АО "Тандер"'):
                rc_clean = rc_clean.replace('АО "Тандер"', '', 1).strip()
            qty = ""
            if "оптимальное" in rest:
                qty = rest.split("оптимальное")[1].strip().split()[0]
            elif "(сгорит" in rest:
                qty = rest.split()[0]
            # Округление до кратного 20 по правилу
            try:
                qty_int = int(qty)
                up = ((qty_int + 19) // 20) * 20
                down = (qty_int // 20) * 20
                if up - qty_int < 10:
                    qty_rounded = up
                else:
                    qty_rounded = down
            except Exception:
                qty_rounded = qty
            # Найти строку в df по названию РЦ
            df_idx = None
            for idx, row in df.iterrows():
                rc_df = row[df.columns[3]]
                rc_df_clean = rc_df.strip()
                if rc_df_clean.startswith('АО "Тандер"'):
                    rc_df_clean = rc_df_clean.replace('АО "Тандер"', '', 1).strip()
                if rc_df_clean == rc_clean:
                    df_idx = idx
                    break
            plan_date = day_tue
            if df_idx is not None and col_sun and col_mon and col_tue:
                row = df.iloc[df_idx]
                try:
                    val_sun = int(row[col_sun])
                    val_mon = int(row[col_mon])
                    val_tue = int(row[col_tue])
                    # Если ничего не сгорит до вторника — вторник
                    # Если сгорит до вторника, но не до понедельника — понедельник
                    # Если сгорит до понедельника — воскресенье
                    if val_tue < val_mon:
                        plan_date = day_mon
                    if val_mon < val_sun:
                        plan_date = day_sun
                except Exception:
                    pass
            row = {
                "Плановая Дата": plan_date.strftime("%d.%m.%Y"),
                "ИНН КА": "3403014273",
                "НАИМЕНОВАНИЕ КА": "АО Сады Придонья",
                "№ Договора": "ГК/35150/18 от 1.08.2018",
                "РЦ Возврата": rc_clean,
                "КОЛ-ВО поддонов к возврату (Фактически отгруженных на РЦ) по данным КА": qty_rounded,
                "КОЛ-ВО поддонов к возврату (Фактически отгруженных на РЦ) по данным АО \"Тандер\"": "",
                "Объем к возврату (кратно 20 шт)": ""
            }
            rows.append(row)
    else:
        day_plus_2 = today + datetime.timedelta(days=2)
        plan_date_str = day_plus_2.strftime("%d.%m")
        for text in filtered:
            rc, rest = text.split(":", 1)
            rc_clean = rc.strip()
            if rc_clean.startswith('АО "Тандер"'):
                rc_clean = rc_clean.replace('АО "Тандер"', '', 1).strip()
            qty = ""
            if "оптимальное" in rest:
                qty = rest.split("оптимальное")[1].strip().split()[0]
            elif "(сгорит" in rest:
                qty = rest.split()[0]
            try:
                qty_int = int(qty)
                up = ((qty_int + 19) // 20) * 20
                down = (qty_int // 20) * 20
                if up - qty_int < 10:
                    qty_rounded = up
                else:
                    qty_rounded = down
            except Exception:
                qty_rounded = qty
            row = {
                "Плановая Дата": day_plus_2.strftime("%d.%m.%Y"),
                "ИНН КА": "3403014273",
                "НАИМЕНОВАНИЕ КА": "АО Сады Придонья",
                "№ Договора": "ГК/35150/18 от 1.08.2018",
                "РЦ Возврата": rc_clean,
                "КОЛ-ВО поддонов к возврату (Фактически отгруженных на РЦ) по данным КА": qty_rounded,
                "КОЛ-ВО поддонов к возврату (Фактически отгруженных на РЦ) по данным АО \"Тандер\"": "",
                "Объем к возврату (кратно 20 шт)": ""
            }
            rows.append(row)
    df_out = pd.DataFrame(rows, columns=columns)
    # Определяем plan_date_str для имени файла
    if rows:
        plan_date_str = rows[0]["Плановая Дата"][:5]  # только день и месяц
    else:
        plan_date_str = today.strftime("%d.%m")
    out_path = f"запрос планирования на {plan_date_str}.xlsx"
    df_out.to_excel(out_path, index=False)

    # --- Устанавливаем ширину столбцов ---
    import openpyxl
    column_widths = [15, 15, 25, 25, 50, 35, 35, 20]  # подберите под ваш шаблон
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    # Устанавливаем высоту первой строки (заголовка)
    ws.row_dimensions[1].height = 35  # например, 30, можно увеличить при необходимости
    # Включаем перенос текста для всех ячеек первой строки (заголовков)
    from openpyxl.styles import Alignment
    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True)
    wb.save(out_path)

    # --- Отправляем файл ---
    await query.message.answer_document(FSInputFile(out_path), caption="Ваш файл сформирован!")
    os.remove(out_path)
    await query.answer()

async def main():
    await dp.start_polling(bot)

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        pass
